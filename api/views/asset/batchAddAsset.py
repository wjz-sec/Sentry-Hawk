import os
from django.core.files.storage import default_storage
from django.http import JsonResponse
from django.views import View
from openpyxl.reader.excel import load_workbook

from SentryHawk import settings
from app.models import Asset_info, Projects, Tags
from django import forms
from api.views.login import clean_form
class BaseAssetBatchAddForm(forms.Form):
    asset = forms.CharField(error_messages={'required': '请输入要添加的资产'}, required=True)
    asset_type = forms.CharField(error_messages={'required': '请输入资产类型'}, required=True)
    asset_project = forms.CharField(error_messages={'required': '请输入资产所属的项目'}, required=True)
    asset_note = forms.CharField(required=False)
# 资产查询数据清洗
class AssetBatchAddForm(BaseAssetBatchAddForm):
    def clean_asset(self):
        asset = self.cleaned_data['asset']
        asset_project = self.data['asset_project']
        if asset_project:
            exists = Projects.objects.filter(name=asset_project).exists()
            if not exists:
                tag = Tags.objects.get(id='4')
                Projects.objects.create(name=asset_project, tag=tag)
            asset_project = Projects.objects.get(name=asset_project)
            if Asset_info.objects.filter(asset=asset, asset_project=asset_project).exists():
                return self.add_error('asset', asset+'资产已存在，请重新输入')
        return asset

    def clean_asset_type(self):
        asset_type = self.cleaned_data['asset_type']
        if asset_type:
            if asset_type not in ['URL', 'IP', 'Domain']:
                return self.add_error('asset_type', '请检查资产类型格式是否正确')
        return asset_type

    def clean_asset_project(self):
        asset_project = self.cleaned_data['asset_project']
        if asset_project:
            asset_project = Projects.objects.filter(name=asset_project).first()
        return asset_project
# 资产批量添加
class BatchAddAsset(View):
    def post(self, request):
        res = {
            'code': 500,
            'msg': "批量添加成功",
            'self': None,
        }

        file = request.FILES.get('file')
        if not file:
            res['msg'] = "上传的内容不是文件,请重新上传"
            return JsonResponse(res)

        # 检查文件扩展名是否为.xlsx
        if not file.name.endswith('.xlsx'):
            res['msg'] = "文件格式不正确，仅支持.xlsx格式"
            return JsonResponse(res)

        # 定义文件保存路径
        tmp_dir = 'tmp'
        file_path = os.path.join(tmp_dir, file.name)

        # 使用 default_storage.save() 方法保存文件
        saved_path = default_storage.save(file_path, file)
        full_path = os.path.join(settings.MEDIA_ROOT, saved_path)

        # 加载工作簿
        try:
            wb = load_workbook(filename=full_path)
            sheet = wb.active  # 或者指定具体的 sheet 名称
            assets_data = []

            # 读取数据（跳过标题行）
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # 假设每行数据是这样的：资产名称，资产类型，数量
                asset, asset_project, asset_type, asset_note = row
                data = {
                    'asset': asset,
                    'asset_type': asset_type,
                    'asset_project': asset_project,
                    'asset_note': asset_note,
                }
                form = AssetBatchAddForm(data)
                if not form.is_valid():
                    res['self'], res['msg'] = clean_form(form)
                    return JsonResponse(res)
                assets_data.append(form.cleaned_data)

            # 批量创建资产
            Asset_info.objects.bulk_create(
                [Asset_info(**data) for data in assets_data]
            )

            wb.close()  # 确保工作簿关闭

        except Exception as e:
            res['msg'] = f"处理文件时出错: {str(e)}"
            return JsonResponse(res)
        finally:
            # 关闭工作簿（以防万一）
            try:
                wb.close()
            except:
                pass
            
            # 删除临时文件前先等待一小段时间
            import time
            time.sleep(0.5)  # 等待500毫秒
            
            try:
                # 尝试删除临时文件
                if os.path.exists(full_path):
                    os.remove(full_path)
            except Exception as e:
                print(f"无法删除临时文件 {full_path}: {str(e)}")
                # 继续执行，不影响返回结果

        res['code'] = 200
        return JsonResponse(res, safe=False, json_dumps_params={'ensure_ascii': False})